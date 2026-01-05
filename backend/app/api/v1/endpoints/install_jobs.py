from __future__ import annotations

import calendar
import uuid
from datetime import datetime, timedelta, timezone
from enum import Enum
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import and_, or_, select
from sqlalchemy.orm import Session

from app.core.audit import log_audit
from app.core.security import require_admin_or_dev, require_view_or_higher
from app.db.session import get_db
from app.models import (
    Certificate,
    CertInstallJob,
    Device,
    JOB_STATUS_CANCELED,
    JOB_STATUS_PENDING,
    JOB_STATUS_REQUESTED,
    User,
    UserDevice,
)
from app.schemas.install_job import InstallJobApproveRequest, InstallJobRead

router = APIRouter(prefix="/install-jobs", tags=["install-jobs"])


class ExportPeriod(str, Enum):
    LAST_15_DAYS = "last_15_days"
    THIS_MONTH = "this_month"
    LAST_6_MONTHS = "last_6_months"


class ExportScope(str, Enum):
    ALL = "all"
    MINE = "mine"
    MY_DEVICE = "my-device"


def subtract_months(value: datetime, months: int) -> datetime:
    month = value.month - months
    year = value.year
    while month <= 0:
        month += 12
        year -= 1
    day = min(value.day, calendar.monthrange(year, month)[1])
    return value.replace(year=year, month=month, day=day)


def resolve_period_range(period: ExportPeriod) -> tuple[datetime, datetime]:
    now = datetime.now(timezone.utc)
    if period == ExportPeriod.THIS_MONTH:
        start = datetime(now.year, now.month, 1, tzinfo=timezone.utc)
    elif period == ExportPeriod.LAST_6_MONTHS:
        start = subtract_months(now, 6)
    else:
        start = now - timedelta(days=15)
    return start, now


def format_datetime(value: datetime | None) -> str:
    if value is None:
        return ""
    return value.isoformat()


@router.get("/export")
async def export_install_jobs(
    period: ExportPeriod = Query(default=ExportPeriod.LAST_15_DAYS),
    scope: ExportScope = Query(default=ExportScope.ALL),
    db: Session = Depends(get_db),
    current_user=Depends(require_view_or_higher),
) -> StreamingResponse:
    start_date, end_date = resolve_period_range(period)
    statement = (
        select(CertInstallJob, Certificate.name, Device.hostname, User.nome, User.ad_username)
        .join(Certificate, Certificate.id == CertInstallJob.cert_id)
        .join(Device, Device.id == CertInstallJob.device_id)
        .join(User, User.id == CertInstallJob.requested_by_user_id)
        .where(
            CertInstallJob.org_id == current_user.org_id,
            CertInstallJob.created_at >= start_date,
            CertInstallJob.created_at <= end_date,
        )
    )

    if scope == ExportScope.ALL:
        if current_user.role_global not in {"ADMIN", "DEV"}:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Forbidden")
    elif scope == ExportScope.MINE:
        statement = statement.where(CertInstallJob.requested_by_user_id == current_user.id)
    elif scope == ExportScope.MY_DEVICE:
        statement = (
            statement.outerjoin(
                UserDevice,
                and_(
                    UserDevice.device_id == CertInstallJob.device_id,
                    UserDevice.user_id == current_user.id,
                ),
            )
            .where(
                or_(
                    CertInstallJob.requested_by_user_id == current_user.id,
                    Device.assigned_user_id == current_user.id,
                    UserDevice.is_allowed.is_(True),
                ),
            )
            .distinct()
        )

    statement = statement.order_by(CertInstallJob.created_at.desc())
    results = db.execute(statement).all()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Jobs"
    headers = [
        "ID",
        "Certificado",
        "Device",
        "Status",
        "Solicitado por",
        "Criado em",
        "Atualizado em",
        "Aprovado em",
        "Iniciado em",
        "Finalizado em",
        "Erro",
    ]
    sheet.append(headers)
    header_fill = PatternFill(fill_type="solid", start_color="E2E8F0", end_color="E2E8F0")
    header_font = Font(bold=True, color="0F172A")
    border = Border(
        left=Side(style="thin", color="CBD5F5"),
        right=Side(style="thin", color="CBD5F5"),
        top=Side(style="thin", color="CBD5F5"),
        bottom=Side(style="thin", color="CBD5F5"),
    )
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for job, cert_name, device_name, requested_name, requested_username in results:
        requester = requested_name or requested_username
        sheet.append(
            [
                str(job.id),
                cert_name,
                device_name,
                job.status,
                requester,
                format_datetime(job.created_at),
                format_datetime(job.updated_at),
                format_datetime(job.approved_at),
                format_datetime(job.started_at),
                format_datetime(job.finished_at),
                job.error_message or "",
            ]
        )
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=sheet.max_column):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[column_letter].width = min(max_length + 3, 50)
    sheet.auto_filter.ref = sheet.dimensions

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    filename = f"jobs_{period.value}.xlsx"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.get("", response_model=list[InstallJobRead])
async def list_install_jobs(
    mine: bool = Query(default=False),
    db: Session = Depends(get_db),
    current_user=Depends(require_view_or_higher),
) -> list[CertInstallJob]:
    if not mine:
        if current_user.role_global not in {"ADMIN", "DEV"}:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Forbidden")
        statement = select(CertInstallJob).where(CertInstallJob.org_id == current_user.org_id)
    else:
        statement = select(CertInstallJob).where(
            CertInstallJob.org_id == current_user.org_id,
            CertInstallJob.requested_by_user_id == current_user.id,
        )
    statement = statement.order_by(CertInstallJob.created_at.desc())
    return db.execute(statement).scalars().all()


@router.get("/mine", response_model=list[InstallJobRead])
async def list_my_jobs(
    db: Session = Depends(get_db), current_user=Depends(require_view_or_higher)
) -> list[CertInstallJob]:
    statement = select(CertInstallJob).where(
        CertInstallJob.org_id == current_user.org_id,
        CertInstallJob.requested_by_user_id == current_user.id,
    )
    statement = statement.order_by(CertInstallJob.created_at.desc())
    return db.execute(statement).scalars().all()


@router.get("/my-device", response_model=list[InstallJobRead])
async def list_my_device_jobs(
    db: Session = Depends(get_db), current_user=Depends(require_view_or_higher)
) -> list[CertInstallJob]:
    statement = (
        select(CertInstallJob)
        .join(Device, Device.id == CertInstallJob.device_id)
        .outerjoin(
            UserDevice,
            and_(
                UserDevice.device_id == CertInstallJob.device_id,
                UserDevice.user_id == current_user.id,
            ),
        )
        .where(
            CertInstallJob.org_id == current_user.org_id,
            or_(
                CertInstallJob.requested_by_user_id == current_user.id,
                Device.assigned_user_id == current_user.id,
                UserDevice.is_allowed.is_(True),
            ),
        )
        .order_by(CertInstallJob.created_at.desc())
        .distinct()
    )
    return db.execute(statement).scalars().all()


@router.post("/{job_id}/approve", response_model=InstallJobRead)
async def approve_job(
    job_id: uuid.UUID,
    payload: InstallJobApproveRequest | None = None,
    db: Session = Depends(get_db),
    current_user=Depends(require_admin_or_dev),
) -> CertInstallJob:
    job = db.get(CertInstallJob, job_id)
    if job is None or job.org_id != current_user.org_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="job not found")
    if job.status != JOB_STATUS_REQUESTED:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="invalid status")

    job.status = JOB_STATUS_PENDING
    job.approved_by_user_id = current_user.id
    job.approved_at = datetime.now(timezone.utc)
    log_audit(
        db=db,
        org_id=current_user.org_id,
        action="INSTALL_APPROVED",
        entity_type="cert_install_job",
        entity_id=job.id,
        actor_user_id=current_user.id,
        meta={"job_id": str(job.id), "reason": payload.reason if payload else None},
    )
    db.commit()
    db.refresh(job)
    return job


@router.post("/{job_id}/deny", response_model=InstallJobRead)
async def deny_job(
    job_id: uuid.UUID,
    payload: InstallJobApproveRequest | None = None,
    db: Session = Depends(get_db),
    current_user=Depends(require_admin_or_dev),
) -> CertInstallJob:
    job = db.get(CertInstallJob, job_id)
    if job is None or job.org_id != current_user.org_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="job not found")
    if job.status != JOB_STATUS_REQUESTED:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="invalid status")

    job.status = JOB_STATUS_CANCELED
    job.approved_by_user_id = current_user.id
    job.approved_at = datetime.now(timezone.utc)
    log_audit(
        db=db,
        org_id=current_user.org_id,
        action="INSTALL_DENIED",
        entity_type="cert_install_job",
        entity_id=job.id,
        actor_user_id=current_user.id,
        meta={"job_id": str(job.id), "reason": payload.reason if payload else None},
    )
    db.commit()
    db.refresh(job)
    return job
