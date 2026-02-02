from __future__ import annotations

import re
import uuid
from datetime import datetime, timedelta, timezone
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.audit import log_audit
from app.core.config import settings
from app.core.security import require_admin_or_dev, require_view_or_higher
from app.db.session import get_db
from app.models import (
    CertInstallJob,
    Certificate,
    CLEANUP_MODE_DEFAULT,
    CLEANUP_MODE_EXEMPT,
    CLEANUP_MODE_KEEP_UNTIL,
    Device,
    JOB_STATUS_PENDING,
    JOB_STATUS_REQUESTED,
    JOB_TYPE_INSTALL,
    UserDevice,
)
from app.schemas.certificate import (
    CertificateCreate,
    CertificatePortalRead,
    CertificateTechnicalRead,
)
from app.schemas.install_job import InstallJobCreate, InstallJobRead

router = APIRouter(prefix="/certificados", tags=["certificados"])

CN_PATTERN = re.compile(r"(?:^|,)\s*CN=([^,]+)", flags=re.IGNORECASE)
CNPJ_PATTERN = re.compile(r"\b\d{2}\.?\d{3}\.?\d{3}/?\d{4}-?\d{2}\b")
CPF_PATTERN = re.compile(r"\b\d{3}\.?\d{3}\.?\d{3}-?\d{2}\b")


def sanitize_certificate_name(value: str) -> str:
    sanitized = value
    patterns = [
        r"senha\s*[:=]?\s*[^\s]+",
        r"senha[_-]?[^\s]+",
        r"\bsenha\b",
    ]
    for pattern in patterns:
        sanitized = re.sub(pattern, "", sanitized, flags=re.IGNORECASE)
    sanitized = re.sub(r"[_-]{2,}", "-", sanitized)
    sanitized = re.sub(r"\s{2,}", " ", sanitized)
    sanitized = re.sub(r"[-_ ]+$", "", sanitized)
    sanitized = re.sub(r"^[-_ ]+", "", sanitized)
    return sanitized.strip()


def _extract_cn(value: str | None) -> str | None:
    if not value:
        return None
    match = CN_PATTERN.search(value)
    return match.group(1).strip() if match else None


def _detect_document(value: str | None) -> tuple[str | None, str | None]:
    if not value:
        return None, None
    for doc_type, pattern in (("CNPJ", CNPJ_PATTERN), ("CPF", CPF_PATTERN)):
        match = pattern.search(value)
        if match:
            digits = re.sub(r"\D", "", match.group(0))
            # Validar o tipo baseado na quantidade de dígitos
            if len(digits) == 14 and doc_type == "CNPJ":
                return doc_type, digits
            if len(digits) == 11 and doc_type == "CPF":
                return doc_type, digits
    # Fallback: extrair dígitos e determinar por comprimento
    digits = re.sub(r"\D", "", value)
    if len(digits) == 14:
        return "CNPJ", digits
    if len(digits) == 11:
        return "CPF", digits
    if len(digits) >= 14:
        return "CNPJ", digits[:14]
    if len(digits) >= 11:
        return "CPF", digits[:11]
    return None, None


def _mask_document(doc_type: str | None, digits: str | None) -> str | None:
    if not doc_type or not digits:
        return None
    if doc_type == "CNPJ" and len(digits) >= 14:
        return f"CNPJ {digits[:2]}{'*' * 8}{digits[10:14]}"
    if doc_type == "CPF" and len(digits) >= 11:
        return f"CPF ***.***.***-{digits[-2:]}"
    return None


def parse_subject_summary(subject: str | None, issuer: str | None) -> dict[str, str | None]:
    cn = _extract_cn(subject)
    issuer_cn = _extract_cn(issuer)
    doc_type, digits = _detect_document(" ".join(filter(None, [subject, cn])))
    return {
        "cn": cn,
        "issuer_cn": issuer_cn,
        "document_type": doc_type,
        "document_masked": _mask_document(doc_type, digits),
        "document_unmasked": digits,
    }


@router.post("", response_model=CertificatePortalRead, status_code=status.HTTP_201_CREATED)
async def create_certificate(
    payload: CertificateCreate,
    db: Session = Depends(get_db),
    current_user=Depends(require_admin_or_dev),
) -> Certificate:
    certificate = Certificate(org_id=current_user.org_id, **payload.model_dump())
    db.add(certificate)
    log_audit(
        db=db,
        org_id=current_user.org_id,
        action="CERT_CREATED",
        entity_type="certificate",
        entity_id=certificate.id,
        actor_user_id=current_user.id,
        meta={"name": certificate.name},
    )
    db.commit()
    db.refresh(certificate)
    summary = parse_subject_summary(certificate.subject, certificate.issuer)
    response = CertificatePortalRead.model_validate(certificate, from_attributes=True)
    return response.model_copy(
        update={"name": sanitize_certificate_name(response.name), **summary}
    )


@router.get("", response_model=list[CertificatePortalRead])
async def list_certificates(
    db: Session = Depends(get_db), current_user=Depends(require_view_or_higher)
) -> list[Certificate]:
    statement = (
        select(Certificate)
        .where(Certificate.org_id == current_user.org_id)
        .order_by(Certificate.created_at)
    )
    certificates = db.execute(statement).scalars().all()
    payload: list[CertificatePortalRead] = []
    for cert in certificates:
        summary = parse_subject_summary(cert.subject, cert.issuer)
        response = CertificatePortalRead.model_validate(cert, from_attributes=True)
        payload.append(
            response.model_copy(
                update={"name": sanitize_certificate_name(response.name), **summary}
            )
        )
    return payload


@router.get("/export/excel")
async def export_certificates(
    db: Session = Depends(get_db), current_user=Depends(require_view_or_higher)
) -> StreamingResponse:
    statement = (
        select(Certificate)
        .where(Certificate.org_id == current_user.org_id)
        .order_by(Certificate.created_at)
    )
    certificates = db.execute(statement).scalars().all()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Certificados"
    headers = [
        "Empresa",
        "Documento",
        "Titular",
        "Serial",
        "SHA1",
        "Validade",
        "Status",
    ]
    sheet.append(headers)
    header_fill = PatternFill(fill_type="solid", start_color="E2E8F0", end_color="E2E8F0")
    header_font = Font(bold=True, color="0F172A")
    border = Border(
        left=Side(style="thin", color="CBD5F5"),
        right=Side(style="thin", color="CBD5F5"),
        top=Side(style="thin", color="CBD5F5"),
        bottom=Side(style="thin", color="CBD5F5"),
    )
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    def get_status(not_after: datetime) -> str:
        now = datetime.now(timezone.utc)
        if not_after <= now:
            return "Vencido"
        days_until = (not_after - now).days
        if days_until <= 7:
            return "Vence em 7d"
        if days_until <= 30:
            return "Vence em 30d"
        return "Válido"

    for cert in certificates:
        summary = parse_subject_summary(cert.subject, cert.issuer)
        sheet.append(
            [
                sanitize_certificate_name(cert.name),
                summary.get("document_masked") or "-",
                summary.get("cn") or "-",
                cert.serial_number or "-",
                cert.sha1_fingerprint or "-",
                cert.not_after.strftime("%d/%m/%Y") if cert.not_after else "-",
                get_status(cert.not_after) if cert.not_after else "-",
            ]
        )

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=sheet.max_column):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border

    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[column_letter].width = min(max_length + 3, 50)

    sheet.auto_filter.ref = sheet.dimensions

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    filename = "certificados.xlsx"
    headers_response = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers_response,
    )


@router.get("/{certificate_id}", response_model=CertificatePortalRead)
async def get_certificate(
    certificate_id: uuid.UUID,
    db: Session = Depends(get_db),
    current_user=Depends(require_view_or_higher),
) -> Certificate:
    certificate = db.get(Certificate, certificate_id)
    if certificate is None or certificate.org_id != current_user.org_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="certificate not found")
    summary = parse_subject_summary(certificate.subject, certificate.issuer)
    response = CertificatePortalRead.model_validate(certificate, from_attributes=True)
    return response.model_copy(
        update={"name": sanitize_certificate_name(response.name), **summary}
    )


@router.get("/{certificate_id}/technical", response_model=CertificateTechnicalRead)
async def get_certificate_technical(
    certificate_id: uuid.UUID,
    db: Session = Depends(get_db),
    current_user=Depends(require_admin_or_dev),
) -> Certificate:
    certificate = db.get(Certificate, certificate_id)
    if certificate is None or certificate.org_id != current_user.org_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="certificate not found")
    response = CertificateTechnicalRead.model_validate(certificate, from_attributes=True)
    return response.model_copy(update={"name": sanitize_certificate_name(response.name)})


@router.post(
    "/{certificate_id}/install",
    response_model=InstallJobRead,
    status_code=status.HTTP_201_CREATED,
)
async def create_install_job(
    certificate_id: uuid.UUID,
    payload: InstallJobCreate,
    db: Session = Depends(get_db),
    current_user=Depends(require_view_or_higher),
) -> CertInstallJob:
    certificate = db.get(Certificate, certificate_id)
    if certificate is None or certificate.org_id != current_user.org_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="certificate not found")

    device = db.get(Device, payload.device_id)
    if device is None or device.org_id != current_user.org_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="device not found")

    # Bloqueio global do device (toggle ADMIN: device.is_allowed)
    if not device.is_allowed:
        log_audit(
            db=db,
            org_id=current_user.org_id,
            action="INSTALL_DENIED",
            entity_type="device",
            entity_id=device.id,
            actor_user_id=current_user.id,
            meta={
                "reason": "device_not_allowed",
                "cert_id": str(certificate.id),
                "device_id": str(device.id),
                "requested_by_user_id": str(current_user.id),
            },
        )
        db.commit()
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="device not allowed")

    if current_user.role_global == "VIEW":
        allowed_device = db.execute(
            select(UserDevice)
            .where(
                UserDevice.device_id == device.id,
                UserDevice.user_id == current_user.id,
                UserDevice.is_allowed.is_(True),
            )
            .limit(1)
        ).scalar_one_or_none()
        if device.assigned_user_id != current_user.id and allowed_device is None:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="device not allowed")

    auto_approved = False
    auto_reason = None
    now = datetime.now(timezone.utc)
    if current_user.role_global in {"DEV", "ADMIN"}:
        initial_status = JOB_STATUS_PENDING
        auto_approved = True
        auto_reason = "role"
    elif current_user.auto_approve_install_jobs is True:
        initial_status = JOB_STATUS_PENDING
        auto_approved = True
        auto_reason = "flag"
    elif device.auto_approve is True:
        initial_status = JOB_STATUS_PENDING
        auto_approved = True
        auto_reason = "device"
    else:
        initial_status = JOB_STATUS_REQUESTED

    cleanup_mode = payload.cleanup_mode or CLEANUP_MODE_DEFAULT
    keep_until = payload.keep_until
    keep_reason = payload.keep_reason
    keep_set_by_user_id = None
    keep_set_at = None

    if cleanup_mode == CLEANUP_MODE_KEEP_UNTIL:
        if not device.allow_keep_until:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="cleanup_mode KEEP_UNTIL not allowed for device",
            )
        if keep_until is None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="keep_until is required when cleanup_mode is KEEP_UNTIL",
            )
        if keep_until.tzinfo is None:
            keep_until = keep_until.replace(tzinfo=timezone.utc)
        if keep_until <= now:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="keep_until must be in the future",
            )
        if current_user.role_global == "VIEW":
            max_until = now + timedelta(hours=settings.retention_keep_until_max_hours)
            if keep_until > max_until:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="keep_until exceeds retention limit for VIEW role",
                )
        keep_set_by_user_id = current_user.id
        keep_set_at = now
    elif cleanup_mode == CLEANUP_MODE_EXEMPT:
        if not device.allow_exempt:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="cleanup_mode EXEMPT not allowed for device",
            )
        if not keep_reason:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="keep_reason is required when cleanup_mode is EXEMPT",
            )
        keep_set_by_user_id = current_user.id
        keep_set_at = now
    else:
        cleanup_mode = CLEANUP_MODE_DEFAULT
        keep_until = None
        keep_reason = None

    job = CertInstallJob(
        org_id=current_user.org_id,
        cert_id=certificate.id,
        device_id=device.id,
        requested_by_user_id=current_user.id,
        status=initial_status,
        job_type=JOB_TYPE_INSTALL,
        cleanup_mode=cleanup_mode,
        keep_until=keep_until,
        keep_reason=keep_reason,
        keep_set_by_user_id=keep_set_by_user_id,
        keep_set_at=keep_set_at,
    )
    if auto_approved:
        job.approved_by_user_id = current_user.id
        job.approved_at = now
    db.add(job)
    db.flush()
    log_audit(
        db=db,
        org_id=current_user.org_id,
        action="INSTALL_REQUESTED",
        entity_type="cert_install_job",
        entity_id=job.id,
        actor_user_id=current_user.id,
        meta={
            "cert_id": str(certificate.id),
            "device_id": str(device.id),
            "status_inicial": initial_status,
            "requested_by_user_id": str(current_user.id),
        },
    )
    if cleanup_mode != CLEANUP_MODE_DEFAULT:
        log_audit(
            db=db,
            org_id=current_user.org_id,
            action="RETENTION_SET",
            entity_type="cert_install_job",
            entity_id=job.id,
            actor_user_id=current_user.id,
            meta={
                "job_id": str(job.id),
                "cert_id": str(certificate.id),
                "device_id": str(device.id),
                "cleanup_mode": cleanup_mode,
                "keep_until": keep_until.isoformat() if keep_until else None,
                "keep_reason": keep_reason,
            },
        )
    if auto_approved:
        log_audit(
            db=db,
            org_id=current_user.org_id,
            action="INSTALL_APPROVED",
            entity_type="cert_install_job",
            entity_id=job.id,
            actor_user_id=current_user.id,
            meta={"auto": True, "via": auto_reason, "job_id": str(job.id)},
        )
    db.commit()
    db.refresh(job)
    return job
